# !/usr/bin python3
# encoding: utf-8 -*-
# @file     : test_001.py
# @author   : 李然
# @Time     : 2025/5/25 9:53
# @Copyright: 焕新生活
import openpyxl
import yaml


def read_excel(filepath,sheet_name):
    # 获取整个文档对象
    wb = openpyxl.load_workbook(filepath)
    print(wb)
    # 获取某个sheet工作表的数据对象
    sheet_data = wb[sheet_name]
    # print(sheet_data)
    lines_count = sheet_data.max_row # 获取总行数
    cols_count = sheet_data.max_column # 获取总列数
    # print(lines_count,cols_count)
    data = [] # 用来存储所有行数据的，每行数据都是这个列表的子列表
    for l in range(2,lines_count+1): # 2,3,4,5
        line = [] # 用来存储当前行所有列的单元格数据的
        for c in range(1,cols_count+1): # 1,2,3,4,5
            cell_data = sheet_data.cell(l,c).value
            line.append(cell_data)
        data.append(line)
    return data

def load_yaml_file(filepath):
    with open(file=filepath,mode='r',encoding='utf-8') as f:
        data = yaml.load(f,Loader=yaml.FullLoader)
        return data

def write_yaml(filepath,content):
    with open(file=filepath,mode='w',encoding='utf-8') as f:
        yaml.dump(content,f,Dumper=yaml.Dumper)

if __name__ == '__main__':
    # print(read_excel(r'D:\pycharmprojects\python20250323\apiobjectframework\data\mtxshop_data.xlsx', '立即购买接口'))
    print(load_yaml_file('../data/coupon_data.yml'))